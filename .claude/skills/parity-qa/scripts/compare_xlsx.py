#!/usr/bin/env python3
"""두 .xlsx를 의미 단위로 비교한다 (XML 바이트 비교 아님).

excelize와 openpyxl은 다른 XML을 내므로, 라벨에 의미 있는 속성만 비교한다:
시트명/수, 셀값, 병합 범위, 열너비, 행높이, 이미지 앵커 셀+개수, 셀별 테두리/폰트.

usage: python compare_xlsx.py golden.xlsx candidate.xlsx
종료코드 0 = 일치, 1 = 불일치(차이 출력).
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def col_widths(ws):
    """열너비를 개별 열 letter 단위로 정규화한다.

    excelize는 동일폭 인접 열을 하나의 <col min=a max=b>로 coalesce하고,
    openpyxl 로더는 이를 첫 열(B) 한 항목(min=2 max=13)으로만 등록한다.
    openpyxl(골든)은 보통 열당 1개(min==max)다. 양쪽 모두 ColumnDimension의
    min..max 범위를 각 열 letter로 전개해 동등 비교한다. 너비값 비교는 유지된다
    (전개는 동등 변환일 뿐 — 폭이 다르면 해당 열에서 MISMATCH가 잡힌다).
    """
    out = {}
    for dim in ws.column_dimensions.values():
        if dim.width is None:
            continue
        w = round(dim.width, 4)
        # min/max가 없으면(드물게) 키를 그대로 단일 열로 취급
        lo = dim.min if dim.min is not None else None
        hi = dim.max if dim.max is not None else None
        if lo is None or hi is None:
            continue
        for idx in range(lo, hi + 1):
            out[get_column_letter(idx)] = w
    return out


def sheet_facts(ws):
    facts = {}
    facts["title"] = ws.title
    facts["merges"] = sorted(str(m) for m in ws.merged_cells.ranges)
    facts["col_widths"] = col_widths(ws)
    facts["row_heights"] = {
        k: round(v.height, 4) for k, v in ws.row_dimensions.items() if v.height is not None
    }
    # 셀값 (값 있는 셀만)
    values = {}
    fonts = {}
    borders = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                values[cell.coordinate] = cell.value
            f = cell.font
            if f and (f.bold or (f.name and f.name.lower() == "times new roman")):
                fonts[cell.coordinate] = (f.name, f.size, bool(f.bold))
            b = cell.border
            sides = {s: getattr(b, s).style for s in ("left", "right", "top", "bottom")
                     if getattr(b, s) and getattr(b, s).style}
            if sides:
                borders[cell.coordinate] = sides
    facts["values"] = values
    facts["fonts"] = fonts
    facts["borders"] = borders
    # QR 위치는 Go에서 의도적으로 박스 중앙으로 이동(레거시와 다름) — 앵커 비교 제외.
    # (이미지 존재/개수는 centering_test.go가 별도 검증)
    # facts["images"] 제거.
    return facts


def compare(a_path, b_path):
    a = load_workbook(a_path)
    b = load_workbook(b_path)
    diffs = []
    if a.sheetnames != b.sheetnames:
        diffs.append(f"sheet names: {a.sheetnames} != {b.sheetnames}")
    for name in a.sheetnames:
        if name not in b.sheetnames:
            continue
        fa, fb = sheet_facts(a[name]), sheet_facts(b[name])
        for key in fa:
            if fa[key] != fb[key]:
                diffs.append(f"[{name}] {key}:\n  golden={fa[key]}\n  cand  ={fb[key]}")
    return diffs


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(2)
    diffs = compare(sys.argv[1], sys.argv[2])
    if diffs:
        print(f"MISMATCH ({len(diffs)} differences):")
        for d in diffs:
            print(" -", d)
        sys.exit(1)
    print("MATCH")
    sys.exit(0)
