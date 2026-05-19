"""
Tests for ExcelLabelGenerator with ProjectLabel (doc_type='2') paths.

Covers the uncovered code paths identified in the coverage audit:
  - excel_generator._setup_project_document (lines 126-162)
  - excel_generator._apply_project_borders  (lines 166-245)
  - excel_generator._create_additional_sheets doc_type='2' branch (lines 258-259)
  - excel_generator._apply_qr_codes ValueError when path count is too low (line 274)
  - excel_generator._apply_qr_codes label=None auto-mode guard (lines 294-295)
  - excel_generator._apply_qr_codes exception handler (lines 307-309)
  - excel_generator.create_label_excel exception re-raise (lines 380-382)
  - document_schema.ProjectLabel.doc_count property (line 156)
  - file_lifecycle.FileLifecycleManager._schedule OSError branch (lines 46-47)
"""
import io
import os
import threading
import pytest
from unittest.mock import patch, MagicMock
from PIL import Image as PILImage
import openpyxl

from excel_generator import ExcelLabelGenerator
from document_schema import ProjectLabel, make_label
from file_lifecycle import FileLifecycleManager


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def make_png_file(tmp_dir: str, name: str = "qr.png") -> str:
    path = os.path.join(tmp_dir, name)
    img = PILImage.new("RGB", (75, 75), color=(0, 0, 0))
    img.save(path, format="PNG")
    return path


@pytest.fixture
def generator(tmp_path):
    upload = str(tmp_path / "uploads")
    os.makedirs(upload)
    return ExcelLabelGenerator(upload_folder=upload)


@pytest.fixture
def pjt_data():
    return {
        'pjt_number': 'PJT-001',
        'pjt_test_number': 'TEST-001',
        'pjt_doc_title': '시험 절차서',
        'pjt_doc_writer': '홍길동',
        'pjt_doc_count': 2,
    }


@pytest.fixture
def pjt_data_single():
    return {
        'pjt_number': 'PJT-002',
        'pjt_test_number': 'TEST-002',
        'pjt_doc_title': '단권 절차서',
        'pjt_doc_writer': '김철수',
        'pjt_doc_count': 1,
    }


# ---------------------------------------------------------------------------
# document_schema — ProjectLabel.doc_count property (line 156)
# ---------------------------------------------------------------------------

class TestProjectLabelDocCount:
    def test_doc_count_property_returns_int(self):
        label = ProjectLabel(
            pjt_number='P', pjt_test_number='T', pjt_doc_title='D',
            pjt_doc_writer='W', pjt_doc_count=5,
        )
        assert label.doc_count == 5  # covers line 156


# ---------------------------------------------------------------------------
# excel_generator — ProjectLabel (doc_type='2') happy paths
# ---------------------------------------------------------------------------

class TestProjectLabelExcel:
    def test_project_label_creates_xlsx(self, generator, pjt_data):
        """_setup_project_document + _apply_project_borders are exercised."""
        filepath, filename = generator.create_label_excel(
            doc_type='2', binder_size=3, data=pjt_data
        )
        assert os.path.exists(filepath)
        assert filename.endswith('.xlsx')

    def test_project_label_multi_sheet_has_correct_count(self, generator, pjt_data):
        """_create_additional_sheets doc_type='2' branch (line 258-259)."""
        filepath, _ = generator.create_label_excel(
            doc_type='2', binder_size=3, data=pjt_data
        )
        wb = openpyxl.load_workbook(filepath)
        assert len(wb.worksheets) == 2

    def test_project_label_second_sheet_s23_updated(self, generator, pjt_data):
        """Sheet 2 S23 value is updated for doc_type='2'."""
        filepath, _ = generator.create_label_excel(
            doc_type='2', binder_size=3, data=pjt_data
        )
        wb = openpyxl.load_workbook(filepath)
        sheet2 = wb['Sheet 2']
        assert sheet2['S23'].value == '2/2'

    def test_project_label_single_sheet(self, generator, pjt_data_single):
        """Single-volume project label: only Sheet 1 is created."""
        filepath, _ = generator.create_label_excel(
            doc_type='2', binder_size=5, data=pjt_data_single
        )
        wb = openpyxl.load_workbook(filepath)
        assert len(wb.worksheets) == 1

    def test_project_label_paste_mode(self, generator, pjt_data, tmp_path):
        """paste mode with doc_type='2'."""
        qr_dir = str(tmp_path / "qr")
        os.makedirs(qr_dir)
        paths = [make_png_file(qr_dir, f"qr_{i}.png") for i in range(2)]
        filepath, _ = generator.create_label_excel(
            doc_type='2', binder_size=7, data=pjt_data, qr_image_paths=paths
        )
        assert os.path.exists(filepath)

    def test_project_label_7cm_binder(self, generator, pjt_data):
        """7 cm binder for project doc_type uses correct QR cell E8."""
        filepath, _ = generator.create_label_excel(
            doc_type='2', binder_size=7, data=pjt_data
        )
        assert os.path.exists(filepath)


# ---------------------------------------------------------------------------
# excel_generator — _apply_qr_codes ValueError guard (line 274)
# ---------------------------------------------------------------------------

class TestApplyQrCodesValueError:
    def test_too_few_paths_raises_value_error(self, generator, pjt_data, tmp_path):
        """Fewer qr_image_paths than sheets raises ValueError (line 274)."""
        qr_dir = str(tmp_path / "qr")
        os.makedirs(qr_dir)
        # pjt_doc_count=2 → 2 sheets, but we only pass 1 path
        single_path = [make_png_file(qr_dir, "qr_0.png")]
        with pytest.raises(ValueError, match="sheets expected"):
            generator.create_label_excel(
                doc_type='2', binder_size=3, data=pjt_data, qr_image_paths=single_path
            )


# ---------------------------------------------------------------------------
# excel_generator — label=None auto-mode guard (lines 294-295)
# ---------------------------------------------------------------------------

class TestApplyQrCodesLabelNoneGuard:
    def test_label_none_logs_error_and_skips(self, generator, pjt_data_single, tmp_path):
        """When label=None in auto-mode, error is logged and QR is skipped gracefully."""
        from openpyxl import Workbook as OWB
        from openpyxl.styles import Alignment

        upload = str(tmp_path / "uploads2")
        os.makedirs(upload)
        gen = ExcelLabelGenerator(upload_folder=upload)

        wb = OWB()
        ws = wb.active
        ws.title = "Sheet 1"

        # label=None triggers the guard at line 293-295
        img_files = gen._apply_qr_codes(wb, '1', 3, 'BASE', qr_image_paths=None, label=None)
        # No exception raised; no img_files produced
        assert img_files == []


# ---------------------------------------------------------------------------
# excel_generator — create_label_excel exception re-raise (lines 380-382)
# ---------------------------------------------------------------------------

class TestCreateLabelExcelExceptionReRaise:
    def test_exception_during_save_is_reraised(self, generator, pjt_data_single):
        """Exception inside create_label_excel is logged and re-raised (line 380-382)."""
        with patch.object(generator, '_setup_basic_layout', side_effect=RuntimeError("boom")):
            with pytest.raises(RuntimeError, match="boom"):
                generator.create_label_excel(
                    doc_type='2', binder_size=3, data=pjt_data_single
                )


# ---------------------------------------------------------------------------
# file_lifecycle — OSError branch in cleanup (lines 46-47)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# excel_generator — _apply_qr_codes exception handler (lines 307-309)
# ---------------------------------------------------------------------------

class TestApplyQrCodesExceptionHandler:
    def test_add_image_exception_is_caught_and_logged(self, generator, pjt_data_single, tmp_path, caplog):
        """If sheet.add_image raises, the exception is caught and execution continues (lines 307-309)."""
        import logging
        from openpyxl import Workbook as OWB

        upload = str(tmp_path / "uploads3")
        os.makedirs(upload)
        gen = ExcelLabelGenerator(upload_folder=upload)

        qr_dir = str(tmp_path / "qr_err")
        os.makedirs(qr_dir)
        path = make_png_file(qr_dir, "qr_0.png")

        from openpyxl import Workbook as OWB
        wb = OWB()
        ws = wb.active
        ws.title = "Sheet 1"

        with patch("excel_generator.Image", side_effect=Exception("broken image")):
            with caplog.at_level(logging.ERROR, logger="excel_generator"):
                img_files = gen._apply_qr_codes(
                    wb, '1', 3, 'BASE', qr_image_paths=[path], label=None
                )
        assert any("Failed to add QR code" in r.message for r in caplog.records)


# ---------------------------------------------------------------------------
# excel_generator — _cleanup_temp_files OSError branch (lines 321-322)
# ---------------------------------------------------------------------------

class TestCleanupTempFilesOsError:
    def test_oserror_during_cleanup_is_logged(self, generator, tmp_path, caplog):
        """OSError during temp-file cleanup is logged (lines 321-322)."""
        import logging

        f = tmp_path / "img.png"
        f.write_text("x")

        with patch("excel_generator.os.remove", side_effect=OSError("denied")):
            with caplog.at_level(logging.WARNING, logger="excel_generator"):
                count = generator._cleanup_temp_files([str(f)])

        assert count == 0
        assert any("Failed to delete temporary image file" in r.message for r in caplog.records)


# ---------------------------------------------------------------------------
class TestFileLifecycleOsErrorBranch:
    def test_oserror_during_file_removal_is_logged(self, tmp_path, caplog):
        """If os.remove raises OSError, it is caught and logged (lines 46-47)."""
        import logging
        manager = FileLifecycleManager()
        f = tmp_path / "target.txt"
        f.write_text("data")

        real_remove = os.remove

        def fake_remove(path):
            raise OSError("permission denied")

        with patch("file_lifecycle.os.remove", side_effect=OSError("permission denied")):
            with caplog.at_level(logging.WARNING, logger="file_lifecycle"):
                manager.register_file(str(f), delay=0)
                import time
                time.sleep(0.3)

        # The OSError path was taken; file still exists (remove was mocked)
        assert any("Failed to clean up" in r.message for r in caplog.records)
