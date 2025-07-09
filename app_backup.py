import logging
from flask import Flask, render_template, request, send_file, redirect, url_for, flash, jsonify, make_response
import os
import qrcode
import threading
import time
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import datetime
import io
from PIL import Image as PILImage
import base64

app = Flask(__name__)
app.secret_key = 'your_secret_key'
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# 로그 설정
LOG_FOLDER = 'logs'
if not os.path.exists(LOG_FOLDER):
    os.makedirs(LOG_FOLDER)
log_file_path = os.path.join(LOG_FOLDER, 'app.log')
logging.basicConfig(filename=log_file_path, level=logging.INFO, format='%(asctime)s %(levelname)s %(name)s %(threadName)s : %(message)s')
logger = logging.getLogger(__name__)

def delete_file_later(filepath, delay=600):
    """Delete a file after a specified delay."""
    def delete_file():
        filename = os.path.basename(filepath)
        logger.info(f"File deletion timer started for {filename} - will delete in {delay} seconds")
        time.sleep(delay)
        if os.path.exists(filepath):
            file_size = os.path.getsize(filepath)
            os.remove(filepath)
            logger.info(f"File deleted successfully: {filename} (was {file_size} bytes)")
        else:
            logger.warning(f"File not found for deletion: {filename}")
    threading.Thread(target=delete_file).start()

@app.route('/')
def index():
    client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
    user_agent = request.environ.get('HTTP_USER_AGENT', 'unknown')
    logger.info(f"Index page accessed from {client_ip} - User-Agent: {user_agent[:100]}{'...' if len(user_agent) > 100 else ''}")
    current_year = datetime.now().year
    return render_template('index.html', current_year=current_year)

def create_qr_code_image(qr_text):
    """QR 코드 이미지를 생성하고 PIL Image 객체를 반환합니다."""
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=2
    )
    qr.add_data(qr_text.encode('CP949'))
    qr.make(fit=False)
    img = qr.make_image(fill_color="black", back_color="white")
    return img

def create_label_excel(doc_type, binder_size, data):
    """라벨 Excel 파일 생성 공통 함수"""
    start_time = time.time()
    logger.info(f"Starting Excel file creation - Doc type: {doc_type}, Binder size: {binder_size}cm")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    img_files = []

    # 기본 라벨 스타일 설정
    ws.row_dimensions[1].height = 2.25
    ws.row_dimensions[2].height = 27
    ws.row_dimensions[3].height = 27
    ws.row_dimensions[4].height = 216
    ws.row_dimensions[5].height = 40.5
    ws.row_dimensions[6].height = 27
    ws.row_dimensions[7].height = 27
    for row in range(8, 18):
        ws.row_dimensions[row].height = 6.75
    ws.row_dimensions[18].height = 2.25

    ws.column_dimensions["A"].width = 0.375
    ws.column_dimensions["N"].width = 0.375

    ws.merge_cells('B2:M2')
    ws.merge_cells('B3:M3')
    ws.merge_cells('B4:M4')
    ws.merge_cells('B5:M5')
    ws.merge_cells('B6:M6')

    thin_box = Border(left=Side(border_style='thin', color='000000'),
                      right=Side(border_style='thin', color='000000'),
                      top=Side(border_style='thin', color='000000'),
                      bottom=Side(border_style='thin', color='000000'))

    for thin_range in ws['B2:M6']:
        for cell in thin_range:
            cell.border = thin_box

    # 경계선 설정
    for thick_range in ws['A1:A18']:
        for cell in thick_range:
            cell.border = Border(left=Side(border_style='medium', color='000000'))
    for thick_range in ws['N1:N18']:
        for cell in thick_range:
            cell.border = Border(right=Side(border_style='medium', color='000000'))
    for thick_range in ws['A1:N1']:
        for cell in thick_range:
            cell.border = Border(top=Side(border_style='medium', color='000000'))
    for thick_range in ws['A18:N18']:
        for cell in thick_range:
            cell.border = Border(bottom=Side(border_style='medium', color='000000'))

    ws['A1'].border = Border(left=Side(border_style='medium', color='000000'),
                             top=Side(border_style='medium', color='000000'))
    ws['N1'].border = Border(right=Side(border_style='medium', color='000000'),
                             top=Side(border_style='medium', color='000000'))
    ws['A18'].border = Border(left=Side(border_style='medium', color='000000'),
                              bottom=Side(border_style='medium', color='000000'))
    ws['N18'].border = Border(right=Side(border_style='medium', color='000000'),
                              bottom=Side(border_style='medium', color='000000'))

    if doc_type == '1':
        # 기기 문서 라벨 생성
        ws.merge_cells('B7:M7')
        for thin_range in ws['B2:M7']:
            for cell in thin_range:
                cell.border = thin_box
        for thin_range in ws['B8:M8']:
            for cell in thin_range:
                cell.border = Border(top=Side(border_style='thin', color='000000'))
        for thin_range in ws['B8:B17']:
            for cell in thin_range:
                cell.border = Border(left=Side(border_style='thin', color='000000'))
        for thin_range in ws['M8:M17']:
            for cell in thin_range:
                cell.border = Border(right=Side(border_style='thin', color='000000'))
        for thin_range in ws['B17:M17']:
            for cell in thin_range:
                cell.border = Border(bottom=Side(border_style='thin', color='000000'))
        ws['B17'].border = Border(left=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
        ws['M17'].border = Border(right=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))

        ws['B2'].value = data['eq_number']
        ws['B2'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['B3'].value = data['eq_doc_number']
        ws['B3'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['B4'].value = data['eq_doc_title']
        ws['B4'].font = Font(name='times new roman', size=16, color='000000', bold=True)
        ws['B6'].value = data['eq_doc_department']
        ws['B6'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['B7'].value = data['eq_doc_year']
        ws['B7'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['B5'] = "{}/{}".format(1, data['eq_doc_count'])
        ws['B5'].font = Font(name='times new roman', size=12, color='000000', bold=True)

        doc_count = int(data['eq_doc_count'])
        base_filename = data['eq_doc_number']
        logger.info(f"Creating {doc_count} sheets for equipment document: {base_filename}")
        
        for i in range(2, doc_count + 1):
            source = wb['Sheet 1']
            destination = wb.copy_worksheet(source)
            destination.title = "Sheet {}".format(i)
            destination["B5"] = "{}/{}".format(i, doc_count)
            for row in destination.rows:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text='true')
            destination["B4"].font = Font(name='times new roman', size=16, color='000000', bold=True)

    else:
        # 과제 문서 라벨 생성
        ws.row_dimensions[20].height = 2.25
        ws.row_dimensions[21].height = 48
        ws.row_dimensions[22].height = 34.5
        ws.row_dimensions[23].height = 27.75
        ws.row_dimensions[24].height = 2.25

        ws.column_dimensions["Q"].width = 8.13
        ws.column_dimensions["R"].width = 34.88
        ws.column_dimensions["S"].width = 8.13
        ws.column_dimensions["T"].width = 0.375

        ws.merge_cells('Q21:S21')
        ws.merge_cells('Q22:S22')

        # 추가 경계선 설정
        for thin_range in ws['B7:M7']:
            for cell in thin_range:
                cell.border = Border(top=Side(border_style='thin', color='000000'))
        for thin_range in ws['B7:B17']:
            for cell in thin_range:
                cell.border = Border(left=Side(border_style='thin', color='000000'))
        for thin_range in ws['M7:M17']:
            for cell in thin_range:
                cell.border = Border(right=Side(border_style='thin', color='000000'))
        for thin_range in ws['B17:M17']:
            for cell in thin_range:
                cell.border = Border(bottom=Side(border_style='thin', color='000000'))
        ws['B17'].border = Border(left=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
        ws['M17'].border = Border(right=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))

        for col in range(14, 17):
            ws.column_dimensions[get_column_letter(col)].width = 0.375

        for thin_range in ws['Q20:S20']:
            for cell in thin_range:
                cell.border = Border(top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))
        for thin_range in ws['Q24:S24']:
            for cell in thin_range:
                cell.border = Border(top=Side(border_style='thin', color='000000'),
                                     bottom=Side(border_style='thin', color='000000'))
        for thin_range in ws['P21:P23']:
            for cell in thin_range:
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'))
        for thin_range in ws['T21:T23']:
            for cell in thin_range:
                cell.border = Border(left=Side(border_style='thin', color='000000'),
                                     right=Side(border_style='thin', color='000000'))
        ws['P20'].border = Border(left=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'))
        ws['T20'].border = Border(right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'))
        ws['T24'].border = Border(right=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))
        ws['P24'].border = Border(left=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))

        for thin_range in ws['Q22:S22']:
            for cell in thin_range:
                cell.border = thin_box

        ws['B2'].value = data['pjt_number']
        ws['B2'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['B3'].value = data['pjt_test_number']
        ws['B3'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['B4'].value = data['pjt_doc_title']
        ws['B4'].font = Font(name='times new roman', size=16, color='000000', bold=True)
        ws['B6'].value = data['pjt_doc_writer']
        ws['B6'].font = Font(name='times new roman', size=12, color='000000', bold=True)

        ws['Q21'].value = "[" + ws['B2'].value + "]" + " " + ws['B3'].value
        ws['Q21'].font = Font(name='times new roman', size=20, bold=True)
        ws['Q21'].alignment = Alignment(horizontal='center', vertical='center', wrap_text='true')

        ws['Q22'].value = ws['B4'].value
        ws['Q22'].font = Font(name='times new roman', size=13, bold=True)
        ws['Q22'].alignment = Alignment(horizontal='center', vertical='center', wrap_text='true')

        ws['R23'].value = ws['B6'].value
        ws['R23'].font = Font(name='times new roman', size=13, bold=True)
        ws['R23'].alignment = Alignment(horizontal='center', vertical='center', wrap_text='true')

        ws['B5'] = "{}/{}".format(1, data['pjt_doc_count'])
        ws['B5'].font = Font(name='times new roman', size=12, color='000000', bold=True)
        ws['S23'] = "{}/{}".format(1, data['pjt_doc_count'])
        ws['S23'].font = Font(name='times new roman', size=12, color='000000', bold=True)

        ws.print_area = 'A1:T24'
        
        doc_count = int(data['pjt_doc_count'])
        base_filename = data['pjt_test_number']
        logger.info(f"Creating {doc_count} sheets for project document: {base_filename}")

        for i in range(2, doc_count + 1):
            source = wb['Sheet 1']
            destination = wb.copy_worksheet(source)
            destination.title = "Sheet {}".format(i)
            destination["B5"] = "{}/{}".format(i, doc_count)
            destination["S23"] = "{}/{}".format(i, doc_count)
            for row in destination.rows:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text='true')
            destination["B4"].font = Font(name='times new roman', size=16, color='000000', bold=True)
            destination.print_area = 'A1:T24'

    for row in ws.rows:
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text='true')

    # QR 코드 생성 및 삽입
    logger.info(f"Starting QR code generation for {len(wb.worksheets)} sheets")
    for sheet in wb.worksheets:
        if doc_type == '1':
            qr_text = sheet["B2"].value + "|" + sheet["B3"].value + "|" + sheet["B4"].value + "|" + sheet["B6"].value + "|" + str(sheet["B7"].value) + "|" + sheet["B5"].value
        else:
            qr_text = sheet["B2"].value + "|" + sheet["B3"].value + "|" + sheet["B4"].value + "|" + sheet["B6"].value + "|" + sheet["B5"].value
        
        qr_img = create_qr_code_image(qr_text)
        img_file = os.path.join(UPLOAD_FOLDER, f"{base_filename}_{sheet.title}.png")
        qr_img.save(img_file)
        img_obj = Image(img_file)

        if binder_size == 7:
            for ws_sheet in wb.worksheets:
                for i in range(ord('B'), ord('N')):
                    column = chr(i)
                    column_dimensions = ws_sheet.column_dimensions[column]
                    column_dimensions.width = 1.875
            cell_pos = "E9" if doc_type == '1' else "E8"
            img_obj.width = 75
            img_obj.height = 75
            sheet.add_image(img_obj, cell_pos)
            img_files.append(img_file)
        elif binder_size == 5:
            for ws_sheet in wb.worksheets:
                for i in range(ord('B'), ord('N')):
                    column = chr(i)
                    column_dimensions = ws_sheet.column_dimensions[column]
                    column_dimensions.width = 1.25
            cell_pos = "D9" if doc_type == '1' else "D8"
            img_obj.width = 75
            img_obj.height = 75
            sheet.add_image(img_obj, cell_pos)
            img_files.append(img_file)
        elif binder_size == 3:
            for ws_sheet in wb.worksheets:
                for i in range(ord('B'), ord('N')):
                    column = chr(i)
                    column_dimensions = ws_sheet.column_dimensions[column]
                    column_dimensions.width = 1
            cell_pos = "D9" if doc_type == '1' else "D8"
            img_obj.width = 75
            img_obj.height = 75
            sheet.add_image(img_obj, cell_pos)
            img_files.append(img_file)
        elif binder_size == 1:
            for ws_sheet in wb.worksheets:
                for i in range(ord('B'), ord('N')):
                    column = chr(i)
                    column_dimensions = ws_sheet.column_dimensions[column]
                    column_dimensions.width = 0.75
            cell_pos = "B9"
            img_obj.width = 75
            img_obj.height = 75
            sheet.add_image(img_obj, cell_pos)
            img_files.append(img_file)

    # 파일 저장
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    filename = f"{base_filename}_{timestamp}.xlsx"
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    logger.info(f"Saving Excel file: {filename}")
    wb.save(filepath)
    wb.close()

    # 임시 이미지 파일 삭제
    deleted_images = 0
    for img_file in img_files:
        try:
            os.remove(img_file)
            deleted_images += 1
        except FileNotFoundError:
            logger.warning(f"Temporary image file not found for deletion: {img_file}")
    
    logger.info(f"Deleted {deleted_images} temporary QR code images")
    
    # 파일 생성 완료 로그
    final_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
    elapsed_time = time.time() - start_time
    logger.info(f"Excel file creation completed - File: {filename}, Size: {final_size} bytes, Time: {elapsed_time:.2f}s, Sheets: {len(wb.worksheets)}")

    return filepath, filename

@app.route('/create_label', methods=['POST'])
def create_label():
    try:
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        user_agent = request.environ.get('HTTP_USER_AGENT', 'unknown')
        logger.info(f"Create label request received from {client_ip} - User-Agent: {user_agent}")
        
        doc_type = request.form.get('doc_type')
        binder_size = int(request.form.get('binder_size'))
        
        logger.info(f"Request details - Document type: {doc_type}, Binder size: {binder_size}cm")
        
        if doc_type == '2' and binder_size == 1:
            flash("과제 문서의 경우 3cm 미만 바인더 크기를 선택할 수 없습니다.", "error")
            logger.warning(f"Invalid binder size selected for project document - Client: {client_ip}, Doc type: {doc_type}, Binder size: {binder_size}cm")
            return redirect(url_for('index'))

        if doc_type == '1':
            required_fields = ['eq_number', 'eq_doc_number', 'eq_doc_title', 'eq_doc_count', 'eq_doc_department', 'eq_doc_year']
        else:
            required_fields = ['pjt_number', 'pjt_test_number', 'pjt_doc_title', 'pjt_doc_writer', 'pjt_doc_count']

        for field in required_fields:
            if not request.form.get(field):
                flash("모든 필드를 채워주세요.", "error")
                logger.warning(f"Missing required field: {field} - Client: {client_ip}, Doc type: {doc_type}")
                return redirect(url_for('index'))

        # 데이터 정리
        data = {}
        if doc_type == '1':
            data['eq_number'] = request.form.get('eq_number', '').strip().replace("\n","").replace("\r","")
            data['eq_doc_number'] = request.form.get('eq_doc_number', '').strip().replace("\n","").replace("\r","")
            data['eq_doc_title'] = request.form.get('eq_doc_title', '').strip().replace("\n","").replace("\r","")
            data['eq_doc_count'] = int(request.form.get('eq_doc_count', '1')) if request.form.get('eq_doc_count', '1').isdigit() else 1
            data['eq_doc_department'] = request.form.get('eq_doc_department', '').strip().replace("\n","").replace("\r","")
            data['eq_doc_year'] = int(request.form.get('eq_doc_year', str(datetime.now().year))) if request.form.get('eq_doc_year', str(datetime.now().year)).isdigit() else datetime.now().year
            
            logger.info(f"Equipment document data - Number: {data['eq_number']}, Doc Number: {data['eq_doc_number']}, Title: {data['eq_doc_title'][:50]}{'...' if len(data['eq_doc_title']) > 50 else ''}, Count: {data['eq_doc_count']}, Department: {data['eq_doc_department']}, Year: {data['eq_doc_year']}")
        else:
            data['pjt_number'] = request.form.get('pjt_number', '').strip().replace("\n","").replace("\r","")
            data['pjt_test_number'] = request.form.get('pjt_test_number', '').strip().replace("\n","").replace("\r","")
            data['pjt_doc_title'] = request.form.get('pjt_doc_title', '').strip().replace("\n","").replace("\r","")
            data['pjt_doc_writer'] = request.form.get('pjt_doc_writer', '').strip().replace("\n","").replace("\r","")
            data['pjt_doc_count'] = int(request.form.get('pjt_doc_count', '1')) if request.form.get('pjt_doc_count', '1').isdigit() else 1
            
            logger.info(f"Project document data - Project: {data['pjt_number']}, Test Number: {data['pjt_test_number']}, Title: {data['pjt_doc_title'][:50]}{'...' if len(data['pjt_doc_title']) > 50 else ''}, Writer: {data['pjt_doc_writer']}, Count: {data['pjt_doc_count']}")

        # 공통 함수 사용하여 라벨 생성
        logger.info(f"Starting label generation process for client {client_ip}")
        filepath, filename = create_label_excel(doc_type, binder_size, data)
        
        # 파일 크기 계산
        file_size = os.path.getsize(filepath) if os.path.exists(filepath) else 0
        logger.info(f"Label generation completed - File: {filename}, Size: {file_size} bytes, Path: {filepath}")
        
        # 파일 다운로드 후 삭제 예약
        delete_file_later(filepath)
        logger.info(f"File deletion scheduled for {filename} (600 seconds delay)")

        # 다운로드 완료 쿠키 설정
        response = make_response(send_file(filepath, as_attachment=True, download_name=filename))
        response.set_cookie('download_complete', 'true', max_age=10)  # 10초 후 자동 삭제
        logger.info(f"File download initiated for client {client_ip} - File: {filename}")
        return response
    except Exception as e:
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        logger.error(f"Error creating label for client {client_ip}: {str(e)}", exc_info=True)
        flash("라벨 생성 중 오류가 발생했습니다.", "error")
        return redirect(url_for('index'))

@app.route('/api/create_label', methods=['POST'])
def api_create_label():
    """라벨 생성 API 엔드포인트"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': '잘못된 JSON 데이터입니다.'}), 400
        
        doc_type = data.get('doc_type')
        binder_size = data.get('binder_size')
        
        # 데이터 유효성 검사
        if doc_type not in ['1', '2']:
            return jsonify({'error': '잘못된 문서 종류입니다.'}), 400
        
        if binder_size not in [1, 3, 5, 7]:
            return jsonify({'error': '잘못된 바인더 크기입니다.'}), 400
        
        if doc_type == '2' and binder_size == 1:
            return jsonify({'error': '과제 문서의 경우 3cm 미만 바인더 크기를 선택할 수 없습니다.'}), 400
        
        # 필수 필드 검사
        if doc_type == '1':
            required_fields = ['eq_number', 'eq_doc_number', 'eq_doc_title', 'eq_doc_count', 'eq_doc_department', 'eq_doc_year']
        else:
            required_fields = ['pjt_number', 'pjt_test_number', 'pjt_doc_title', 'pjt_doc_writer', 'pjt_doc_count']
        
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'필수 필드가 누락되었습니다: {field}'}), 400
        
        # 라벨 생성
        filepath, filename = create_label_excel(doc_type, binder_size, data)
        
        # 파일 다운로드 후 삭제 예약
        delete_file_later(filepath)
        
        # 성공 응답
        return jsonify({
            'success': True,
            'message': '라벨이 성공적으로 생성되었습니다.',
            'filename': filename,
            'download_url': f'/download/{filename}'
        })
        
    except Exception as e:
        logger.error(f"API Error creating label: {e}")
        return jsonify({'error': '라벨 생성 중 오류가 발생했습니다.'}), 500

@app.route('/api/qr_image/<path:qr_text>', methods=['GET'])
def api_qr_image(qr_text):
    """주어진 텍스트로 QR 코드 이미지를 PNG 형식으로 반환합니다."""
    try:
        if not qr_text:
            return jsonify({'error': 'QR 코드 텍스트가 제공되지 않았습니다.'}), 400
        
        # QR 코드 이미지 생성
        qr_img = create_qr_code_image(qr_text)
        
        # 메모리에서 이미지 생성
        img_io = io.BytesIO()
        qr_img.save(img_io, 'PNG')
        img_io.seek(0)
        
        return send_file(
            img_io,
            mimetype='image/png',
            as_attachment=False,
            download_name=f'qr_code.png'
        )
        
    except Exception as e:
        logger.error(f"Error generating QR code image: {e}")
        return jsonify({'error': 'QR 코드 이미지 생성 중 오류가 발생했습니다.'}), 500

@app.route('/api/qr_image_base64', methods=['POST'])
def api_qr_image_base64():
    """주어진 텍스트로 QR 코드 이미지를 base64 형식으로 반환합니다."""
    try:
        data = request.get_json()
        if not data or not data.get('text'):
            return jsonify({'error': 'QR 코드 텍스트가 제공되지 않았습니다.'}), 400
        
        qr_text = data['text']
        
        # QR 코드 이미지 생성
        qr_img = create_qr_code_image(qr_text)
        
        # base64로 변환
        img_io = io.BytesIO()
        qr_img.save(img_io, 'PNG')
        img_io.seek(0)
        
        img_base64 = base64.b64encode(img_io.getvalue()).decode('utf-8')
        
        return jsonify({
            'success': True,
            'image_base64': img_base64,
            'mime_type': 'image/png'
        })
        
    except Exception as e:
        logger.error(f"Error generating QR code base64: {e}")
        return jsonify({'error': 'QR 코드 이미지 생성 중 오류가 발생했습니다.'}), 500

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """파일 다운로드 엔드포인트"""
    try:
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        if os.path.exists(filepath):
            return send_file(filepath, as_attachment=True, download_name=filename)
        else:
            return jsonify({'error': '파일을 찾을 수 없습니다.'}), 404
    except Exception as e:
        logger.error(f"Error downloading file: {e}")
        return jsonify({'error': '파일 다운로드 중 오류가 발생했습니다.'}), 500

@app.route('/logs')
def logs_page():
    """로그 확인 웹 페이지"""
    client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
    logger.info(f"Logs page accessed from {client_ip}")
    return render_template('logs.html')

@app.route('/api/logs', methods=['GET'])
def api_get_logs():
    """로그 조회 API 엔드포인트"""
    try:
        client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        # 쿼리 파라미터 처리
        lines = request.args.get('lines', 100, type=int)
        level = request.args.get('level', 'all').upper()
        search = request.args.get('search', '')
        
        logger.info(f"API logs request from {client_ip} - lines: {lines}, level: {level}, search: '{search}'")
        
        # 최대 1000줄로 제한
        lines = min(lines, 1000)
        
        log_file_path = os.path.join(LOG_FOLDER, 'app.log')
        
        if not os.path.exists(log_file_path):
            return jsonify({
                'success': True,
                'logs': [],
                'message': '로그 파일이 아직 생성되지 않았습니다.'
            })
        
        logs = []
        try:
            with open(log_file_path, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                
            # 최근 줄부터 가져오기
            recent_lines = all_lines[-lines:] if len(all_lines) > lines else all_lines
            
            for line in recent_lines:
                line = line.strip()
                if not line:
                    continue
                
                # 레벨 필터링
                if level != 'ALL':
                    if level not in line:
                        continue
                
                # 검색어 필터링
                if search and search.lower() not in line.lower():
                    continue
                
                logs.append(line)
                
        except Exception as e:
            logger.error(f"Error reading log file: {e}")
            return jsonify({'error': '로그 파일을 읽는 중 오류가 발생했습니다.'}), 500
        
        return jsonify({
            'success': True,
            'logs': logs,
            'total_lines': len(logs),
            'requested_lines': lines,
            'level_filter': level,
            'search_filter': search
        })
        
    except Exception as e:
        logger.error(f"Error in logs API: {e}")
        return jsonify({'error': '로그 조회 중 오류가 발생했습니다.'}), 500

@app.route('/api/logs/clear', methods=['POST'])
def api_clear_logs():
    """로그 파일 초기화 API"""
    try:
        log_file_path = os.path.join(LOG_FOLDER, 'app.log')
        
        if os.path.exists(log_file_path):
            # 로그 파일 백업
            backup_path = f"app_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            try:
                with open(log_file_path, 'r', encoding='utf-8') as src:
                    with open(backup_path, 'w', encoding='utf-8') as dst:
                        dst.write(src.read())
                logger.info(f"Log file backed up to: {backup_path}")
            except Exception as backup_error:
                logger.warning(f"Failed to backup log file: {backup_error}")
            
            # 로그 파일 초기화
            open(log_file_path, 'w').close()
            logger.info("Log file cleared by user request")
            
            return jsonify({
                'success': True,
                'message': '로그 파일이 초기화되었습니다.',
                'backup_file': backup_path if 'backup_path' in locals() else None
            })
        else:
            return jsonify({
                'success': True,
                'message': '초기화할 로그 파일이 없습니다.'
            })
            
    except Exception as e:
        logger.error(f"Error clearing logs: {e}")
        return jsonify({'error': '로그 초기화 중 오류가 발생했습니다.'}), 500

@app.route('/api/logs/download', methods=['GET'])
def api_download_logs():
    """로그 파일 다운로드 API"""
    try:
        log_file_path = os.path.join(LOG_FOLDER, 'app.log')
        
        if not os.path.exists(log_file_path):
            return jsonify({'error': '다운로드할 로그 파일이 없습니다.'}), 404
        
        # 다운로드용 파일명 생성
        download_name = f"app_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        
        return send_file(
            log_file_path,
            as_attachment=True,
            download_name=download_name,
            mimetype='text/plain'
        )
        
    except Exception as e:
        logger.error(f"Error downloading logs: {e}")
        return jsonify({'error': '로그 다운로드 중 오류가 발생했습니다.'}), 500

@app.route('/api/health', methods=['GET'])
def api_health():
    """서비스 상태 확인 엔드포인트"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

@app.route('/api/docs')
def api_docs():
    """API 문서 페이지"""
    client_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
    logger.info(f"API documentation page accessed from {client_ip}")
    return render_template('api_docs.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)