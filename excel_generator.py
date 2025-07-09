"""
Excel label generation utilities
"""
import os
import time
import logging
from datetime import datetime
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

from qr_generator import default_qr_generator
from utils import get_file_size_safe, generate_timestamp_filename

logger = logging.getLogger(__name__)

class ExcelLabelGenerator:
    """Excel 라벨 생성을 담당하는 클래스"""
    
    def __init__(self, upload_folder):
        self.upload_folder = upload_folder
        self.qr_generator = default_qr_generator
        
        # 스타일 상수 정의
        self.FONT_TIMES = Font(name='times new roman', size=12, color='000000', bold=True)
        self.FONT_TITLE = Font(name='times new roman', size=16, color='000000', bold=True)
        self.THIN_BORDER = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        self.MEDIUM_BORDER = Border(
            left=Side(border_style='medium', color='000000'),
            right=Side(border_style='medium', color='000000'),
            top=Side(border_style='medium', color='000000'),
            bottom=Side(border_style='medium', color='000000')
        )
    
    def _setup_basic_layout(self, ws):
        """기본 라벨 레이아웃을 설정합니다."""
        # 행 높이 설정
        row_heights = {
            1: 2.25, 2: 27, 3: 27, 4: 216, 5: 40.5, 6: 27, 7: 27
        }
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height
        
        # 8-17행 높이 설정
        for row in range(8, 18):
            ws.row_dimensions[row].height = 6.75
        ws.row_dimensions[18].height = 2.25
        
        # 열 너비 설정
        ws.column_dimensions["A"].width = 0.375
        ws.column_dimensions["N"].width = 0.375
        
        # 셀 병합
        merge_ranges = ['B2:M2', 'B3:M3', 'B4:M4', 'B5:M5', 'B6:M6']
        for range_str in merge_ranges:
            ws.merge_cells(range_str)
    
    def _apply_borders(self, ws, doc_type):
        """테두리를 적용합니다."""
        # 기본 테두리 적용
        for range_str in ['B2:M6']:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = self.THIN_BORDER
        
        # 외곽 굵은 테두리
        border_ranges = [
            ('A1:A18', 'left'), ('N1:N18', 'right'),
            ('A1:N1', 'top'), ('A18:N18', 'bottom')
        ]
        
        for range_str, side in border_ranges:
            border_style = {side: Side(border_style='medium', color='000000')}
            for row in ws[range_str]:
                for cell in row:
                    cell.border = Border(**border_style)
        
        # 모서리 특별 처리
        corners = [
            ('A1', ['left', 'top']), ('N1', ['right', 'top']),
            ('A18', ['left', 'bottom']), ('N18', ['right', 'bottom'])
        ]
        
        for cell_addr, sides in corners:
            border_dict = {side: Side(border_style='medium', color='000000') for side in sides}
            ws[cell_addr].border = Border(**border_dict)
    
    def _setup_equipment_document(self, ws, data):
        """기기 문서 레이아웃을 설정합니다."""
        # 추가 병합
        ws.merge_cells('B7:M7')
        
        # 추가 테두리 설정
        additional_borders = ['B2:M7', 'B8:M8', 'B8:B17', 'M8:M17', 'B17:M17']
        for range_str in additional_borders:
            for row in ws[range_str]:
                for cell in row:
                    if range_str == 'B8:M8':
                        cell.border = Border(top=Side(border_style='thin', color='000000'))
                    elif range_str in ['B8:B17', 'M8:M17']:
                        side = 'left' if 'B' in range_str else 'right'
                        cell.border = Border(**{side: Side(border_style='thin', color='000000')})
                    elif range_str == 'B17:M17':
                        cell.border = Border(bottom=Side(border_style='thin', color='000000'))
                    else:
                        cell.border = self.THIN_BORDER
        
        # 데이터 입력
        ws['B2'].value = data['eq_number']
        ws['B2'].font = self.FONT_TIMES
        ws['B3'].value = data['eq_doc_number']
        ws['B3'].font = self.FONT_TIMES
        ws['B4'].value = data['eq_doc_title']
        ws['B4'].font = self.FONT_TITLE
        ws['B6'].value = data['eq_doc_department']
        ws['B6'].font = self.FONT_TIMES
        ws['B7'].value = data['eq_doc_year']
        ws['B7'].font = self.FONT_TIMES
        ws['B5'].value = f"1/{data['eq_doc_count']}"
        ws['B5'].font = self.FONT_TIMES
        
        return data['eq_doc_number'], int(data['eq_doc_count'])
    
    def _setup_project_document(self, ws, data):
        """과제 문서 레이아웃을 설정합니다."""
        # 추가 행 설정
        additional_rows = {20: 2.25, 21: 48, 22: 34.5, 23: 27.75, 24: 2.25}
        for row, height in additional_rows.items():
            ws.row_dimensions[row].height = height
        
        # 추가 열 설정
        additional_columns = {
            "Q": 8.13, "R": 34.88, "S": 8.13, "T": 0.375
        }
        for col, width in additional_columns.items():
            ws.column_dimensions[col].width = width
        
        # 추가 병합
        ws.merge_cells('Q21:S21')
        ws.merge_cells('Q22:S22')
        
        # 과제 문서용 추가 테두리 설정
        self._apply_project_borders(ws)
        
        # 데이터 입력
        ws['B2'].value = data['pjt_number']
        ws['B2'].font = self.FONT_TIMES
        ws['B3'].value = data['pjt_test_number']
        ws['B3'].font = self.FONT_TIMES
        ws['B4'].value = data['pjt_doc_title']
        ws['B4'].font = self.FONT_TITLE
        ws['B6'].value = data['pjt_doc_writer']
        ws['B6'].font = self.FONT_TIMES
        ws['B5'].value = f"1/{data['pjt_doc_count']}"
        ws['B5'].font = self.FONT_TIMES
        
        # 우측 섹션 데이터
        ws['Q21'].value = f"[{ws['B2'].value}] {ws['B3'].value}"
        ws['Q21'].font = Font(name='times new roman', size=20, bold=True)
        ws['Q21'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws['Q22'].value = ws['B4'].value
        ws['Q22'].font = Font(name='times new roman', size=13, bold=True)
        ws['Q22'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws['R23'].value = ws['B6'].value
        ws['R23'].font = Font(name='times new roman', size=13, bold=True)
        ws['R23'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        ws['S23'].value = f"1/{data['pjt_doc_count']}"
        ws['S23'].font = self.FONT_TIMES
        
        ws.print_area = 'A1:T24'
        
        return data['pjt_test_number'], int(data['pjt_doc_count'])
    
    def _apply_project_borders(self, ws):
        """과제 문서용 추가 테두리를 적용합니다."""
        from openpyxl.utils import get_column_letter
        
        # B7:M17 영역 테두리 설정
        for range_str in ['B7:M7']:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = Border(top=Side(border_style='thin', color='000000'))
        
        for range_str in ['B7:B17']:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = Border(left=Side(border_style='thin', color='000000'))
        
        for range_str in ['M7:M17']:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = Border(right=Side(border_style='thin', color='000000'))
        
        for range_str in ['B17:M17']:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = Border(bottom=Side(border_style='thin', color='000000'))
        
        # 모서리 셀 특별 처리
        ws['B17'].border = Border(
            left=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        ws['M17'].border = Border(
            right=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        
        # N, O, P 열 너비 설정
        for col in range(14, 17):  # N=14, O=15, P=16
            ws.column_dimensions[get_column_letter(col)].width = 0.375
        
        # Q20:S20, Q24:S24 테두리
        for range_str in ['Q20:S20', 'Q24:S24']:
            for row in ws[range_str]:
                for cell in row:
                    cell.border = Border(
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000')
                    )
        
        # P21:P23, T21:T23 테두리
        for range_str in ['P21:P23', 'T21:T23']:
            for row in ws[range_str]:
                for cell in row:
                    side_style = 'left' if 'P' in range_str else 'right'
                    if side_style == 'left':
                        cell.border = Border(
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000')
                        )
                    else:
                        cell.border = Border(
                            left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000')
                        )
        
        # 우측 영역 모서리 셀 특별 처리
        corner_borders = {
            'P20': ['left', 'top'],
            'T20': ['right', 'top'],
            'P24': ['left', 'bottom'],
            'T24': ['right', 'bottom']
        }
        
        for cell_addr, sides in corner_borders.items():
            border_dict = {}
            for side in sides:
                border_dict[side] = Side(border_style='thin', color='000000')
            ws[cell_addr].border = Border(**border_dict)
        
        # Q22:S22 영역 내부 테두리
        for row in ws['Q22:S22']:
            for cell in row:
                cell.border = self.THIN_BORDER
    
    def _create_additional_sheets(self, wb, doc_type, doc_count, base_filename):
        """추가 시트를 생성합니다."""
        logger.info(f"Creating {doc_count} sheets for document: {base_filename}")
        
        for i in range(2, doc_count + 1):
            source = wb['Sheet 1']
            destination = wb.copy_worksheet(source)
            destination.title = f"Sheet {i}"
            destination["B5"].value = f"{i}/{doc_count}"
            
            if doc_type == '2':  # 과제 문서
                destination["S23"].value = f"{i}/{doc_count}"
                destination.print_area = 'A1:T24'
            
            # 정렬 설정 적용
            for row in destination.rows:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            destination["B4"].font = self.FONT_TITLE
    
    def _apply_qr_codes(self, wb, doc_type, binder_size, base_filename):
        """QR 코드를 생성하고 시트에 추가합니다."""
        logger.info(f"Starting QR code generation for {len(wb.worksheets)} sheets")
        img_files = []
        
        # 바인더 크기별 설정
        binder_configs = {
            7: {'column_width': 1.875, 'cell_pos': 'E9' if doc_type == '1' else 'E8'},
            5: {'column_width': 1.25, 'cell_pos': 'D9' if doc_type == '1' else 'D8'},
            3: {'column_width': 1, 'cell_pos': 'D9' if doc_type == '1' else 'D8'},
            1: {'column_width': 0.75, 'cell_pos': 'B9'}
        }
        
        config = binder_configs.get(binder_size, binder_configs[3])
        
        # 모든 시트의 열 너비 조정
        for ws_sheet in wb.worksheets:
            for i in range(ord('B'), ord('N')):
                column = chr(i)
                ws_sheet.column_dimensions[column].width = config['column_width']
        
        # 각 시트에 QR 코드 추가
        for sheet in wb.worksheets:
            try:
                # QR 텍스트 생성
                if doc_type == '1':
                    qr_text = "|".join([
                        str(sheet["B2"].value), str(sheet["B3"].value),
                        str(sheet["B4"].value), str(sheet["B6"].value),
                        str(sheet["B7"].value), str(sheet["B5"].value)
                    ])
                else:
                    qr_text = "|".join([
                        str(sheet["B2"].value), str(sheet["B3"].value),
                        str(sheet["B4"].value), str(sheet["B6"].value),
                        str(sheet["B5"].value)
                    ])
                
                # QR 이미지 생성
                img_file = self.qr_generator.create_qr_for_excel(
                    qr_text, self.upload_folder, f"{base_filename}_{sheet.title}"
                )
                img_files.append(img_file)
                
                # Excel에 이미지 추가
                img_obj = Image(img_file)
                img_obj.width = 75
                img_obj.height = 75
                sheet.add_image(img_obj, config['cell_pos'])
                
            except Exception as e:
                logger.error(f"Failed to add QR code to sheet {sheet.title}: {e}")
                continue
        
        return img_files
    
    def _cleanup_temp_files(self, img_files):
        """임시 이미지 파일들을 정리합니다."""
        deleted_images = 0
        for img_file in img_files:
            try:
                if os.path.exists(img_file):
                    os.remove(img_file)
                    deleted_images += 1
            except OSError as e:
                logger.warning(f"Failed to delete temporary image file {img_file}: {e}")
        
        logger.info(f"Deleted {deleted_images} temporary QR code images")
        return deleted_images
    
    def create_label_excel(self, doc_type, binder_size, data):
        """라벨 Excel 파일을 생성합니다."""
        start_time = time.time()
        logger.info(f"Starting Excel file creation - Doc type: {doc_type}, Binder size: {binder_size}cm")
        
        try:
            # 워크북 생성
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet 1"
            
            # 기본 레이아웃 설정
            self._setup_basic_layout(ws)
            self._apply_borders(ws, doc_type)
            
            # 문서 타입별 설정
            if doc_type == '1':
                base_filename, doc_count = self._setup_equipment_document(ws, data)
            else:
                base_filename, doc_count = self._setup_project_document(ws, data)
            
            # 추가 시트 생성
            if doc_count > 1:
                self._create_additional_sheets(wb, doc_type, doc_count, base_filename)
            
            # 정렬 설정 적용
            for row in ws.rows:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # QR 코드 생성 및 추가
            img_files = self._apply_qr_codes(wb, doc_type, binder_size, base_filename)
            
            # 파일 저장
            filename = generate_timestamp_filename(base_filename)
            filepath = os.path.join(self.upload_folder, filename)
            logger.info(f"Saving Excel file: {filename}")
            
            wb.save(filepath)
            wb.close()
            
            # 임시 파일 정리
            self._cleanup_temp_files(img_files)
            
            # 완료 로그
            final_size = get_file_size_safe(filepath)
            elapsed_time = time.time() - start_time
            logger.info(f"Excel file creation completed - File: {filename}, Size: {final_size} bytes, "
                       f"Time: {elapsed_time:.2f}s, Sheets: {len(wb.worksheets)}")
            
            return filepath, filename
            
        except Exception as e:
            logger.error(f"Failed to create Excel file: {e}", exc_info=True)
            raise